# - Passar as informações de planilha para um arquivo word 
# - Salvar aquele arquivo word em uma pasta especifica (contratos)
# - Repetir para todas as linhas da planilha

from openpyxl import load_workbook
from docx import Document
from datetime import datetime


planilha_fornecedores = load_workbook('./fornecedores.xlsx')
pagina_fornecedores = planilha_fornecedores['Sheet1']

for linha in pagina_fornecedores.iter_rows(min_row=2, values_only= True):
    

